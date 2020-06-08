import  openpyxl
# 1.创建一个Book对象
book=openpyxl.Workbook()
# 2.创建时会自动创建一个sheet，通过active获取
sheet=book.active
# 3.通过sheet对象进行修改
sheet.title='工资表' #修改sheet标题


sheet1=book.create_sheet('年龄表',1)
sheet2=book.create_sheet('体重表',0)

#4.根据名称获取某个表单对象
# ageSht=book['年龄表']
# 5.根据行号和列号给每个单元格赋值，注意和xlrd的不同，是从1开始

sheet1['A1']='name'#写标题栏
sheet1['B1']='age'

sheet1.cell(1,1,value='19')
book.save('new.xls') # 保存文件
