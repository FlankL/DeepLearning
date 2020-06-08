import openpyxl
# 导入颜色样式
from openpyxl.styles import Font,colors
# 导入图片
from openpyxl.drawing.image import  Image

book=openpyxl.load_workbook('TestExcel.xlsx')
sh1=book['2017']
# １．插入行，列
sh1.insert_rows(2)#在第二行的位置插入一行
sh1.insert_cols(2)
# 遍历工作表
for row  in sh1.values,sh1.iter_rows():
    print(row)
# 获取行的编号
for row in sh1.iter_rows(min_row=1,max_col=3,max_row=12):
    # print(row)
    pass
# 2.修改文字颜色，字体，大小等样式风格，都是通过Ｆont对象设定的
sh1['A1'].font=Font(color="F12300",size=15,bold=True,italic=True)
font=openpyxl.styles.Font(color="FFBC00",size=15,bold=True,italic=True)
# 指定整行的风格
for col in range(1,10):
    sh1.cell(row=3,column=col).font=font
    sh1.cell(row=col,column=3).font=font

# 3.插入图片
sh1.add_image(Image('1.png'),'D1')#在第一行第二列的位置插入图片
book.save('update.xlsx') 